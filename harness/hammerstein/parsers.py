"""Multi-format plain-text extraction for --context-file (ham-025 v1)."""

from __future__ import annotations

import csv
from pathlib import Path

FILE_READ_MAX_CHARS = 16384

_TRUNCATION_MARKER = "\n\n[...content truncated at 16K chars...]"

SUPPORTED_EXTENSIONS = frozenset({".pdf", ".docx", ".xlsx", ".csv", ".md", ".txt"})


def _sanitize_utf8(text: str) -> str:
    return text.encode("utf-8", errors="replace").decode("utf-8")


def _maybe_truncate(text: str) -> str:
    if len(text) <= FILE_READ_MAX_CHARS:
        return text
    return text[:FILE_READ_MAX_CHARS] + _TRUNCATION_MARKER


def _markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    norm: list[list[str]] = []
    for row in rows:
        cells = ["" if c is None else str(c) for c in row]
        # Drop fully empty trailing rows common in spreadsheets
        norm.append(cells)
    while norm and not any(c.strip() for c in norm[-1]):
        norm.pop()
    if not norm:
        return ""
    width = max(len(r) for r in norm)
    header = norm[0] + [""] * (width - len(norm[0]))
    lines: list[str] = []
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * width) + " |")
    for body_row in norm[1:]:
        padded = body_row + [""] * (width - len(body_row))
        lines.append("| " + " | ".join(padded) + " |")
    return "\n".join(lines)


def _extract_md_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    chunks: list[str] = []
    for page in reader.pages:
        t = page.extract_text()
        chunks.append(t if t else "")
    return "\n\n".join(chunks)


def _extract_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    parts: list[str] = []
    try:
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append(["" if c is None else str(c) for c in row])
            parts.append(f"## {sheetname}\n")
            parts.append(_markdown_table(rows))
    finally:
        wb.close()
    return "\n\n".join(parts)


def _extract_csv(path: Path) -> str:
    rows: list[list[str]] = []
    with path.open(newline="", encoding="utf-8", errors="replace") as handle:
        for row in csv.reader(handle):
            rows.append(list(row))
    return _markdown_table(rows)


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"unsupported file extension: {path.suffix}")

    if suffix in (".md", ".txt"):
        raw = _extract_md_txt(path)
    elif suffix == ".pdf":
        raw = _extract_pdf(path)
    elif suffix == ".docx":
        raw = _extract_docx(path)
    elif suffix == ".xlsx":
        raw = _extract_xlsx(path)
    elif suffix == ".csv":
        raw = _extract_csv(path)

    sanitized = _sanitize_utf8(raw)
    return _maybe_truncate(sanitized)
